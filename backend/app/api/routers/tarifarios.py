from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io
from datetime import date, timedelta


from app.database import get_db
from app.models.tarifario import Tarifario
from app.models.prestador import Prestador
from app.schemas.tarifario import TarifarioCreate, TarifarioUpdate, TarifarioResponse

router = APIRouter(prefix="/tarifarios", tags=["tarifarios"])

TIPOS_VALIDOS = {
    "correctivo",
    "pre_correctivo",
    "instalacion_desinstalacion",
    "preventivo",
    "guardia",
    "sistemas",
}

_TIPO_ALIASES = {
    "correctivo": "correctivo",
    "pre-correctivo": "pre_correctivo",
    "pre correctivo": "pre_correctivo",
    "instalación-desinstalación": "instalacion_desinstalacion",
    "instalacion-desinstalacion": "instalacion_desinstalacion",
    "instalación desinstalación": "instalacion_desinstalacion",
    "instalacion desinstalacion": "instalacion_desinstalacion",
    "preventivo": "preventivo",
    "guardia": "guardia",
    "sistemas": "sistemas",
}


def _normalize_tipo(raw: str) -> Optional[str]:
    t = raw.lower().strip()
    if t in TIPOS_VALIDOS:
        return t
    if t in _TIPO_ALIASES:
        return _TIPO_ALIASES[t]
    if "pre" in t and "correctiv" in t:
        return "pre_correctivo"
    if "correctiv" in t:
        return "correctivo"
    if "instal" in t or "desinstal" in t:
        return "instalacion_desinstalacion"
    if "prevent" in t:
        return "preventivo"
    if "guardia" in t:
        return "guardia"
    if "sistema" in t:
        return "sistemas"
    return None


def _str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("nan", "none", "") else None


def _rebuild_temporal_chain(db: Session, prestador_id: int, tipo_servicio: str, zona: Optional[str] = None):
    tarifas = (
        db.query(Tarifario)
        .filter(
            Tarifario.prestador_id == prestador_id,
            Tarifario.tipo_servicio == tipo_servicio,
            Tarifario.zona == zona,
        )
        .order_by(Tarifario.vigencia_desde.asc())
        .all()
    )
    for i in range(len(tarifas) - 1):
        tarifas[i].vigencia_hasta = tarifas[i+1].vigencia_desde - timedelta(days=1)
    if tarifas:
        # Asegurar que el último no se pise si ya tenía algo distinto, pero si no tiene nada queda None.
        # En caso de importar nuevo más reciente, el anterior se cierra y el nuevo queda abierto (None).
        # Si el usuario agregó uno en el medio, se cerrará con el siguiente.
        pass


def _build_plantilla() -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Tarifarios"

    header_font = Font(bold=True)
    label_fill = PatternFill("solid", fgColor="D9E1F2")

    ws["A1"] = "PRESTADOR:"
    ws["A1"].font = header_font
    ws["A1"].fill = label_fill
    ws["B1"] = "NOMBRE_CORTO_AQUI"
    ws["B1"].font = Font(color="C00000", bold=True)

    ws["A2"] = "VIGENCIA DESDE:"
    ws["A2"].font = header_font
    ws["A2"].fill = label_fill
    ws["B2"] = "2026-01-01"
    ws["B2"].font = Font(color="C00000", bold=True)

    ws["A3"] = (
        "PRESTADOR: nombre_corto exacto del PST.  "
        "VIGENCIA DESDE: fecha en formato YYYY-MM-DD.  "
        "tipo_servicio debe ser uno de los valores de la fila de nota abajo."
    )
    ws["A3"].font = Font(italic=True, color="808080", size=9)
    ws.merge_cells("A3:F3")

    # Encabezados
    cols = ["tipo_servicio", "zona", "costo_servicio", "costo_km", "vigencia_hasta"]
    col_fill = PatternFill("solid", fgColor="4472C4")
    col_font = Font(bold=True, color="FFFFFF")
    border = Border(
        bottom=Border(bottom=None).bottom,
        right=Border(right=None).right,
    )
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=4, column=i, value=col)
        c.font = col_font
        c.fill = col_fill
        c.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    examples = [
        ["correctivo",                  "",             58042, 649.53, ""],
        ["pre_correctivo",              "",             58042, 649.53, ""],
        ["instalacion_desinstalacion",  "",             58042, 649.53, ""],
        ["preventivo",                  "",             26832,      0, ""],
        ["correctivo",       "Zona Patagonia",          72000, 800,    ""],
    ]
    ex_fill = PatternFill("solid", fgColor="EBF3FB")
    for r, row_data in enumerate(examples, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = ex_fill
            cell.font = Font(italic=True, color="595959")

    # Nota de tipos válidos
    last_row = 5 + len(examples)
    ws.cell(row=last_row, column=1,
            value="► Tipos válidos: correctivo | pre_correctivo | instalacion_desinstalacion | preventivo | guardia | sistemas")
    ws.cell(row=last_row, column=1).font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(f"A{last_row}:F{last_row}")

    ws.cell(row=last_row + 1, column=1,
            value="► zona: dejar vacío si el tarifario aplica a toda la cobertura del PST. vigencia_hasta: vacío = vigente.")
    ws.cell(row=last_row + 1, column=1).font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(f"A{last_row + 1}:F{last_row + 1}")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _find_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=20):
        vals = [str(c.value).lower().strip() for c in row if c.value is not None]
        if "tipo_servicio" in vals:
            return row[0].row
    return None


# ─── Rutas ────────────────────────────────────────────────────────────────────

@router.get("/plantilla")
def descargar_plantilla():
    buf = _build_plantilla()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_tarifarios.xlsx"},
    )


@router.post("/importar-excel")
async def importar_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    data = await file.read()

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    ws = wb.active

    # 1. Leer PRESTADOR y VIGENCIA del encabezado
    nombre_corto = _str(ws["B1"].value)
    if not nombre_corto or nombre_corto.upper() == "NOMBRE_CORTO_AQUI":
        raise HTTPException(status_code=422, detail="Completá el campo PRESTADOR (celda B1) con el nombre_corto del PST.")

    vigencia_raw = _str(ws["B2"].value)
    if not vigencia_raw:
        raise HTTPException(status_code=422, detail="Completá el campo VIGENCIA DESDE (celda B2) en formato YYYY-MM-DD.")
    try:
        vigencia_desde = date.fromisoformat(str(vigencia_raw).strip()[:10])
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Formato de fecha inválido en B2: '{vigencia_raw}'. Usá YYYY-MM-DD.")

    prestador = db.query(Prestador).filter(Prestador.nombre_corto == nombre_corto).first()
    if not prestador:
        prestador = db.query(Prestador).filter(Prestador.nombre_corto.ilike(nombre_corto)).first()
    if not prestador:
        raise HTTPException(status_code=404, detail=f"No existe un prestador con nombre_corto '{nombre_corto}'.")

    prestador_id = prestador.id

    # 2. Encontrar fila de encabezados de columnas
    header_row = _find_header_row(ws)
    if header_row is None:
        raise HTTPException(status_code=422, detail="No se encontró la fila de encabezados (debe contener 'tipo_servicio').")

    headers = {
        str(ws.cell(header_row, c).value).lower().strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value
    }

    def col_idx(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    ci_tipo        = col_idx("tipo_servicio", "tipo")
    ci_zona        = col_idx("zona")
    ci_costo_serv  = col_idx("costo_servicio", "costo serv", "costo_serv")
    ci_costo_km    = col_idx("costo_km", "costo km")
    ci_hasta       = col_idx("vigencia_hasta", "hasta")

    if not ci_tipo or not ci_costo_serv:
        raise HTTPException(
            status_code=422,
            detail=f"Columnas requeridas no encontradas. Se necesitan: tipo_servicio, costo_servicio. Encontradas: {list(headers.keys())}"
        )

    # 3. Tarifarios existentes (para dedup)
    existing = {
        (t.tipo_servicio, t.vigencia_desde)
        for t in db.query(Tarifario).filter(Tarifario.prestador_id == prestador_id).all()
    }

    creados = omitidos = errores = 0
    rebuild_keys = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_tipo = _str(row[ci_tipo - 1]) if ci_tipo else None
        if not raw_tipo:
            continue
        if raw_tipo.startswith(("►", "#", "//", "*")):
            continue

        tipo = _normalize_tipo(raw_tipo)
        if not tipo:
            errores += 1
            continue

        try:
            costo_serv = float(row[ci_costo_serv - 1])
        except (TypeError, ValueError):
            errores += 1
            continue

        costo_km = 0.0
        if ci_costo_km:
            try:
                v = row[ci_costo_km - 1]
                if v is not None:
                    costo_km = float(v)
            except (TypeError, ValueError):
                pass

        zona = _str(row[ci_zona - 1]) if ci_zona else None

        vigencia_hasta = None
        if ci_hasta:
            raw_hasta = _str(row[ci_hasta - 1])
            if raw_hasta:
                try:
                    vigencia_hasta = date.fromisoformat(str(raw_hasta)[:10])
                except ValueError:
                    pass

        key = (tipo, vigencia_desde)
        if key in existing:
            omitidos += 1
            continue

        try:
            t = Tarifario(
                prestador_id=prestador_id,
                tipo_servicio=tipo,
                zona=zona,
                costo_servicio=costo_serv,
                costo_km=costo_km,
                vigencia_desde=vigencia_desde,
                vigencia_hasta=vigencia_hasta,
            )
            db.add(t)
            existing.add(key)
            creados += 1
            rebuild_keys.add((prestador_id, tipo, zona))
        except Exception:
            errores += 1

    db.flush()
    for pid, t_serv, z in rebuild_keys:
        _rebuild_temporal_chain(db, pid, t_serv, z)
    db.commit()
    return {
        "prestador": nombre_corto,
        "prestador_id": prestador_id,
        "vigencia_desde": vigencia_desde.isoformat(),
        "creados": creados,
        "omitidos": omitidos,
        "errores": errores,
    }


@router.get("/", response_model=List[TarifarioResponse])
def list_tarifarios(prestador_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(Tarifario)
    if prestador_id:
        q = q.filter(Tarifario.prestador_id == prestador_id)
    return q.order_by(Tarifario.prestador_id, Tarifario.vigencia_desde.desc()).all()


@router.post("/", response_model=TarifarioResponse, status_code=201)
def create_tarifario(data: TarifarioCreate, db: Session = Depends(get_db)):
    t = Tarifario(**data.model_dump())
    db.add(t)
    db.flush()
    _rebuild_temporal_chain(db, t.prestador_id, t.tipo_servicio, t.zona)
    db.commit()
    db.refresh(t)
    return t


@router.get("/{id}", response_model=TarifarioResponse)
def get_tarifario(id: int, db: Session = Depends(get_db)):
    t = db.query(Tarifario).filter(Tarifario.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Tarifario no encontrado")
    return t


@router.put("/{id}", response_model=TarifarioResponse)
def update_tarifario(id: int, data: TarifarioUpdate, db: Session = Depends(get_db)):
    t = db.query(Tarifario).filter(Tarifario.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Tarifario no encontrado")
    
    old_prestador_id = t.prestador_id
    old_tipo = t.tipo_servicio
    old_zona = t.zona

    for k, v in data.model_dump(exclude_none=True).items():
        setattr(t, k, v)
        
    db.flush()
    _rebuild_temporal_chain(db, old_prestador_id, old_tipo, old_zona)
    if (t.prestador_id != old_prestador_id or t.tipo_servicio != old_tipo or t.zona != old_zona):
        _rebuild_temporal_chain(db, t.prestador_id, t.tipo_servicio, t.zona)
        
    db.commit()
    db.refresh(t)
    return t


@router.delete("/{id}", status_code=204)
def delete_tarifario(id: int, db: Session = Depends(get_db)):
    t = db.query(Tarifario).filter(Tarifario.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Tarifario no encontrado")
    
    prestador_id = t.prestador_id
    tipo_servicio = t.tipo_servicio
    zona = t.zona

    db.delete(t)
    db.flush()
    _rebuild_temporal_chain(db, prestador_id, tipo_servicio, zona)
    db.commit()

