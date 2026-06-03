from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import math
import io

from app.database import get_db
from app.models.tabla_km import TablaKM
from app.models.prestador import Prestador
from app.models.spst import SPST
from app.schemas.tabla_km import TablaKMCreate, TablaKMUpdate, TablaKMResponse

router = APIRouter(prefix="/tabla-km", tags=["tabla_km"])

UMBRAL_DEFAULT = 30.0


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _nan(v) -> bool:
    return v is None or (isinstance(v, float) and math.isnan(v))


def _str(v) -> Optional[str]:
    if _nan(v):
        return None
    s = str(v).strip()
    return s if s and s.lower() != "nan" else None


def _col(df, *names):
    lower = {c.lower(): c for c in df.columns}
    for n in names:
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def _build_plantilla() -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla KM"

    # ── Cabecera de configuración ──
    header_font = Font(bold=True)
    label_fill = PatternFill("solid", fgColor="D9E1F2")
    ws["A1"] = "PRESTADOR:"
    ws["A1"].font = header_font
    ws["A1"].fill = label_fill
    ws["B1"] = "NOMBRE_CORTO_AQUI"
    ws["B1"].font = Font(color="C00000", bold=True)

    ws["A2"] = "UMBRAL VIÁTICO (km):"
    ws["A2"].font = header_font
    ws["A2"].fill = label_fill
    ws["B2"] = 30

    ws["A3"] = "Completar PRESTADOR con el nombre_corto exacto del PST. El umbral viático es 30 km por defecto."
    ws["A3"].font = Font(italic=True, color="808080", size=9)
    ws.merge_cells("A3:H3")

    # ── Encabezados de columnas ──
    cols = ["empresa", "sucursal", "domicilio", "localidad", "provincia", "base", "kms", "url_maps"]
    col_fill = PatternFill("solid", fgColor="4472C4")
    col_font = Font(bold=True, color="FFFFFF")
    border = Border(
        bottom=Side(style="thin", color="2F5496"),
        right=Side(style="thin", color="2F5496"),
    )
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=4, column=i, value=col)
        c.font = col_font
        c.fill = col_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    # ── Fila de ejemplo ──
    example = ["OCA", "VILLA MERCEDES", "Av. Mitre 123", "Villa Mercedes", "San Luis", "Nombre SPST", 45.5, ""]
    ex_fill = PatternFill("solid", fgColor="EBF3FB")
    for i, val in enumerate(example, 1):
        c = ws.cell(row=5, column=i, value=val)
        c.fill = ex_fill
        c.font = Font(italic=True, color="595959")

    # ── Notas ──
    ws["A6"] = "► empresa y sucursal son obligatorios. base = nombre del SPST (se crea automáticamente si no existe). kms = número."
    ws["A6"].font = Font(italic=True, color="808080", size=9)
    ws.merge_cells("A6:H6")

    # Anchos
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_prestador_from_sheet(ws) -> tuple[Optional[str], float]:
    """Reads PRESTADOR from B1 and umbral from B2."""
    nombre_corto = _str(ws["B1"].value)
    try:
        umbral = float(ws["B2"].value)
    except (TypeError, ValueError):
        umbral = UMBRAL_DEFAULT
    return nombre_corto, umbral


def _find_header_row(ws) -> Optional[int]:
    """Returns the 1-based row number of the header row (contains 'empresa')."""
    for row in ws.iter_rows(min_row=1, max_row=20):
        vals = [str(c.value).lower() for c in row if c.value is not None]
        if "empresa" in vals:
            return row[0].row
    return None


# ─── Rutas ────────────────────────────────────────────────────────────────────

@router.get("/plantilla")
def descargar_plantilla():
    buf = _build_plantilla()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_tabla_km.xlsx"},
    )


@router.post("/importar-excel")
async def importar_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    data = await file.read()

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {e}")

    ws = wb.active

    # 1. Leer prestador y umbral del encabezado
    nombre_corto, umbral = _parse_prestador_from_sheet(ws)
    if not nombre_corto or nombre_corto.upper() == "NOMBRE_CORTO_AQUI":
        raise HTTPException(status_code=422, detail="Completá el campo PRESTADOR (celda B1) con el nombre_corto del PST.")

    prestador = db.query(Prestador).filter(Prestador.nombre_corto == nombre_corto).first()
    if not prestador:
        # Buscar case-insensitive
        prestador = db.query(Prestador).filter(
            Prestador.nombre_corto.ilike(nombre_corto)
        ).first()
    if not prestador:
        raise HTTPException(status_code=404, detail=f"No existe un prestador con nombre_corto '{nombre_corto}'. Verificá el campo B1.")

    prestador_id = prestador.id

    # 2. Encontrar fila de encabezados de columnas
    header_row = _find_header_row(ws)
    if header_row is None:
        raise HTTPException(status_code=422, detail="No se encontró la fila de encabezados (debe contener 'empresa').")

    # Mapear columnas por nombre
    headers = {str(ws.cell(header_row, c).value).lower().strip(): c for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value}

    def col_idx(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    ci_empresa   = col_idx("empresa")
    ci_sucursal  = col_idx("sucursal")
    ci_kms       = col_idx("kms", "kms recorrido", "km")
    ci_base      = col_idx("base", "spst", "prestador")
    ci_domicilio = col_idx("domicilio")
    ci_localidad = col_idx("localidad")
    ci_provincia = col_idx("provincia")
    ci_url       = col_idx("url_maps", "url", "maps", "recorrido")

    if not ci_empresa or not ci_sucursal or not ci_kms:
        raise HTTPException(
            status_code=422,
            detail=f"Columnas requeridas no encontradas. Se necesitan: empresa, sucursal, kms. Encontradas: {list(headers.keys())}"
        )

    # 3. Cargar SPSTs existentes
    spst_map = {s.nombre.lower(): s for s in db.query(SPST).filter(SPST.prestador_id == prestador_id).all()}

    # 4. Cargar claves existentes de tabla_km
    existing_keys = {
        (e.lower(), s.lower())
        for e, s in db.query(TablaKM.empresa_nombre, TablaKM.sucursal_nombre)
        .filter(TablaKM.prestador_id == prestador_id).all()
    }

    importadas = omitidas = errores = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Saltar filas de ejemplo / notas (detectar por contenido no numérico en kms)
        raw_empresa  = _str(row[ci_empresa - 1])  if ci_empresa  else None
        raw_sucursal = _str(row[ci_sucursal - 1]) if ci_sucursal else None
        if not raw_empresa or not raw_sucursal:
            continue
        # Saltar si parece fila de nota (empresa empieza con ► o similar)
        if raw_empresa.startswith(("►", "#", "//", "*")):
            continue

        try:
            kms = float(row[ci_kms - 1])
            if _nan(kms):
                continue
        except (TypeError, ValueError):
            continue

        key = (raw_empresa.lower(), raw_sucursal.lower())
        if key in existing_keys:
            omitidas += 1
            continue

        # Resolver o crear SPST
        raw_base = _str(row[ci_base - 1]) if ci_base else None
        spst_id = None
        if raw_base:
            spst_key = raw_base.lower()
            if spst_key not in spst_map:
                new_spst = SPST(prestador_id=prestador_id, nombre=raw_base)
                db.add(new_spst)
                db.flush()
                spst_map[spst_key] = new_spst
            spst_id = spst_map[spst_key].id

        raw_dom  = _str(row[ci_domicilio - 1]) if ci_domicilio else None
        raw_loc  = _str(row[ci_localidad - 1]) if ci_localidad else None
        raw_prov = _str(row[ci_provincia - 1]) if ci_provincia else None
        raw_url  = _str(row[ci_url - 1])       if ci_url       else None

        try:
            t = TablaKM(
                prestador_id=prestador_id,
                spst_id=spst_id,
                empresa_nombre=raw_empresa,
                sucursal_nombre=raw_sucursal,
                domicilio_cliente=raw_dom,
                localidad_cliente=raw_loc,
                provincia_cliente=raw_prov,
                kms_recorrido=kms,
                umbral_viatico=umbral,
                url_maps=raw_url if raw_url and raw_url.startswith("http") else None,
            )
            t.aplica_viatico = t.kms_recorrido > t.umbral_viatico
            t.kms_a_facturar = t.kms_recorrido if t.aplica_viatico else 0.0
            db.add(t)
            existing_keys.add(key)
            importadas += 1
        except Exception:
            errores += 1

    db.commit()
    return {
        "prestador": nombre_corto,
        "prestador_id": prestador_id,
        "importadas": importadas,
        "omitidas": omitidas,
        "errores": errores,
    }


@router.get("/", response_model=List[TablaKMResponse])
def list_tabla_km(
    prestador_id: Optional[int] = None,
    empresa: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(TablaKM)
    if prestador_id:
        q = q.filter(TablaKM.prestador_id == prestador_id)
    if empresa:
        q = q.filter(TablaKM.empresa_nombre.ilike(f"%{empresa}%"))
    return q.order_by(TablaKM.empresa_nombre, TablaKM.sucursal_nombre).all()


@router.post("/", response_model=TablaKMResponse, status_code=201)
def create_tabla_km(data: TablaKMCreate, db: Session = Depends(get_db)):
    t = TablaKM(**data.model_dump())
    t.aplica_viatico = t.kms_recorrido > t.umbral_viatico
    t.kms_a_facturar = t.kms_recorrido if t.aplica_viatico else 0.0
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@router.get("/{id}", response_model=TablaKMResponse)
def get_tabla_km(id: int, db: Session = Depends(get_db)):
    t = db.query(TablaKM).filter(TablaKM.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    return t


@router.put("/{id}", response_model=TablaKMResponse)
def update_tabla_km(id: int, data: TablaKMUpdate, db: Session = Depends(get_db)):
    t = db.query(TablaKM).filter(TablaKM.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(t, k, v)
    t.aplica_viatico = t.kms_recorrido > t.umbral_viatico
    t.kms_a_facturar = t.kms_recorrido if t.aplica_viatico else 0.0
    db.commit()
    db.refresh(t)
    return t


@router.delete("/{id}", status_code=204)
def delete_tabla_km(id: int, db: Session = Depends(get_db)):
    t = db.query(TablaKM).filter(TablaKM.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Entrada no encontrada")
    db.delete(t)
    db.commit()
